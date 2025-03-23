"""
=========================================
@Project ：AITestCase
@Date ：2025/3/22 上午9:39
@Author:Echoxiawan
@Comment:Excel导出模块，负责将测试用例导出为Excel文件
=========================================

"""
import os
from typing import Any, Dict, List, Optional, Tuple, Union

import pandas as pd
from datetime import datetime
import numpy as np
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import OUTPUT_DIR, DEFAULT_EXCEL_FILENAME
from utils.logger import get_logger
from utils.helpers import sanitize_filename, ensure_dir_exists, extract_domain

# 获取日志记录器
logger = get_logger(__name__)

class ExcelExporter:
    """Excel导出器，负责将测试用例导出为Excel文件"""
    
    def __init__(self, output_dir: Optional[str] = None):
        """
        初始化Excel导出器
        
        Args:
            output_dir (Optional[str]): 输出目录，如果为None则使用配置文件中的目录
        """
        self.output_dir = output_dir or OUTPUT_DIR
        # 确保输出目录存在
        ensure_dir_exists(self.output_dir)
        logger.info(f"Excel导出器初始化完成，输出目录: {self.output_dir}")
        
    def export_test_cases(
        self, 
        test_cases: List[Dict[str, Any]], 
        filename: Optional[str] = None, 
        include_timestamp: bool = True,
        metadata: Optional[Dict[str, Any]] = None
    ) -> str:
        """
        导出测试用例为Excel文件
        
        Args:
            test_cases (List[Dict[str, Any]]): 测试用例列表
            filename (Optional[str]): 输出文件名，如果为None则使用默认文件名
            include_timestamp (bool): 是否在文件名中包含时间戳
            metadata (Optional[Dict[str, Any]]): 元数据，包含URLs和需求文档信息等
            
        Returns:
            str: 输出文件的完整路径
        """
        if not test_cases:
            logger.warning("没有测试用例可导出")
            return ""
            
        try:
            # 处理文件名
            if filename is None:
                filename = DEFAULT_EXCEL_FILENAME
                
            # 清理文件名
            clean_filename = sanitize_filename(filename)
            
            # 如果需要添加时间戳
            if include_timestamp:
                # 获取当前时间戳
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                base_name, ext = os.path.splitext(clean_filename)
                clean_filename = f"{base_name}_{timestamp}{ext}"
            
            # 构建完整输出路径
            output_path = os.path.join(self.output_dir, clean_filename)
            
            # 将测试用例转换为DataFrame
            df, main_content_df, page_dfs, requirement_dfs = self._test_cases_to_dataframe(test_cases)
            
            # 导出为Excel
            self._export_dataframe_to_excel(
                df, 
                output_path, 
                main_content_df=main_content_df,
                page_dfs=page_dfs,
                requirement_dfs=requirement_dfs,
                metadata=metadata
            )
            
            logger.info(f"已将 {len(test_cases)} 个测试用例导出到 {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"导出测试用例时出错: {str(e)}")
            return ""
            
    def _test_cases_to_dataframe(self, test_cases: List[Dict[str, Any]]) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, pd.DataFrame], Dict[str, pd.DataFrame]]:
        """
        将测试用例列表转换为DataFrame
        
        Args:
            test_cases (List[Dict[str, Any]]): 测试用例列表
            
        Returns:
            Tuple[pd.DataFrame, pd.DataFrame, Dict[str, pd.DataFrame], Dict[str, pd.DataFrame]]: 
                包含全部测试用例的DataFrame,
                仅包含主内容区域测试用例的DataFrame,
                按页面分组的测试用例DataFrame字典,
                按需求文档分组的测试用例DataFrame字典
        """
        # 标准化字段名称（统一大小写和命名风格）
        normalized_test_cases = []
        
        # 页面和需求文档分组
        page_groups = {}
        requirement_groups = {}
        
        # 要在内部逻辑中使用但不导出的字段列表
        internal_fields = ["测试区域", "需求来源", "test_area"]
        
        for tc in test_cases:
            normalized_tc = {}
            
            # 标准化常见字段
            field_mapping = {
                # 原始字段名: 标准化字段名
                "test_id": "测试ID",
                "testid": "测试ID",
                "id": "测试ID",
                "test_title": "测试标题",
                "title": "测试标题",
                "name": "测试标题",
                "description": "测试标题",
                "priority": "优先级",
                "test_priority": "优先级",
                "preconditions": "前置条件",
                "precondition": "前置条件",
                "prerequisites": "前置条件",
                "test_steps": "测试步骤",
                "steps": "测试步骤",
                "step": "测试步骤",
                "expected_results": "预期结果",
                "expected_result": "预期结果",
                "expected": "预期结果",
                "test_data": "测试数据",
                "data": "测试数据",
                "test_type": "测试类型",
                "type": "测试类型",
                "category": "测试类型",
                "page_source": "页面来源",
                "status": "测试状态",
                "test_status": "测试状态",
                "state": "测试状态"
            }
            
            # 分析测试区域（仅用于内部分组，不导出）
            test_area = "主内容区域"  # 默认假设是主内容区域
            
            # 检查测试标题和步骤中是否包含导航相关词汇
            navigation_keywords = ["导航", "菜单", "导航栏", "menu", "navigation", "nav", "header", "footer", 
                                  "页眉", "页脚", "侧边栏", "sidebar", "顶部栏", "底部栏", "链接跳转", "路由", 
                                  "route", "标签页", "tab", "登录", "注册", "login", "register"]
            
            title = tc.get("test_title", "") or tc.get("title", "") or tc.get("name", "") or ""
            steps = tc.get("test_steps", []) or tc.get("steps", []) or []
            
            # 如果步骤是字符串，转为列表便于检查
            if isinstance(steps, str):
                steps = [steps]
                
            # 检查标题
            if any(keyword in title.lower() for keyword in navigation_keywords):
                test_area = "页面框架"
                
            # 检查步骤
            steps_text = " ".join(str(step) for step in steps)
            if any(keyword in steps_text.lower() for keyword in navigation_keywords):
                test_area = "页面框架"
                
            # 如果测试用例自带了测试区域标记，使用该标记
            if "test_area" in tc:
                test_area = tc["test_area"]
            elif "area" in tc:
                test_area = tc["area"]
                
            # 仅用于内部使用的测试区域字段，不会导出
            normalized_tc["_测试区域"] = test_area
            
            # 提取页面来源和需求来源
            page_source = tc.get("page_source", "未知页面")
            requirement_source = tc.get("requirement_source", "未知需求")
            
            # 如果有页面来源，添加到页面分组中
            if page_source not in page_groups:
                page_groups[page_source] = []
            
            # 如果有需求来源，添加到需求分组中
            if requirement_source not in requirement_groups:
                requirement_groups[requirement_source] = []
                
            # 应用字段映射
            for orig_key, norm_key in field_mapping.items():
                if orig_key in tc and norm_key not in internal_fields:
                    # 处理列表字段，如测试步骤和预期结果
                    if norm_key in ["测试步骤", "预期结果"] and isinstance(tc[orig_key], list):
                        # 格式化为编号列表
                        formatted_list = []
                        for i, item in enumerate(tc[orig_key]):
                            formatted_list.append(f"{i+1}. {item}")
                        normalized_tc[norm_key] = "\n".join(formatted_list)
                    else:
                        normalized_tc[norm_key] = tc[orig_key]
            
            # 确保所有必需字段都存在
            required_fields = ["测试ID", "测试标题", "优先级", "前置条件", "测试步骤", "预期结果"]
            
            for field in required_fields:
                if field not in normalized_tc and field not in internal_fields:
                    if field == "测试ID":
                        # 生成一个唯一ID
                        normalized_tc[field] = f"TC{len(normalized_test_cases) + 1:03d}"
                    elif field == "优先级":
                        normalized_tc[field] = "中"
                    elif field in ["前置条件", "测试数据"]:
                        normalized_tc[field] = "无"
                    else:
                        normalized_tc[field] = ""
            
            # 将原始测试用例中的其他字段也添加到标准化测试用例中
            for key, value in tc.items():
                if key not in field_mapping and key not in field_mapping.values() and key not in internal_fields:
                    # 只添加尚未添加的字段
                    reverse_mapping = {v: k for k, v in field_mapping.items()}
                    if key not in reverse_mapping and key not in internal_fields:
                        normalized_tc[key] = value
            
            normalized_test_cases.append(normalized_tc)
            
            # 添加到页面分组和需求分组
            page_groups[page_source].append(normalized_tc)
            requirement_groups[requirement_source].append(normalized_tc)
            
        # 创建主DataFrame
        df = pd.DataFrame(normalized_test_cases)
        
        # 创建主内容区域DataFrame（使用内部字段进行筛选，但不包含在输出中）
        if "_测试区域" in df.columns:
            main_content_df = df[df["_测试区域"] == "主内容区域"].copy()
            
            # 删除内部字段，不导出
            if "_测试区域" in main_content_df.columns:
                main_content_df = main_content_df.drop(columns=["_测试区域"])
        else:
            main_content_df = pd.DataFrame()
        
        # 从主DataFrame中删除内部字段，不导出
        if "_测试区域" in df.columns:
            df = df.drop(columns=["_测试区域"])
        
        # 创建页面分组DataFrame
        page_dfs = {}
        for page, cases in page_groups.items():
            if cases:  # 确认有用例
                page_df = pd.DataFrame(cases)
                # 删除内部字段
                if "_测试区域" in page_df.columns:
                    page_df = page_df.drop(columns=["_测试区域"])
                page_dfs[page] = page_df
                
        # 创建需求分组DataFrame
        requirement_dfs = {}
        for req, cases in requirement_groups.items():
            if cases:  # 确认有用例
                req_df = pd.DataFrame(cases)
                # 删除内部字段
                if "_测试区域" in req_df.columns:
                    req_df = req_df.drop(columns=["_测试区域"])
                requirement_dfs[req] = req_df
                
        # 记录统计信息
        logger.info(f"测试用例总数: {len(normalized_test_cases)}")
        logger.info(f"主内容区域测试用例数: {len(main_content_df)}")
        logger.info(f"页面分组数: {len(page_dfs)}")
        logger.info(f"需求分组数: {len(requirement_dfs)}")
        
        return df, main_content_df, page_dfs, requirement_dfs
            
    def _export_dataframe_to_excel(
            self, 
            df: pd.DataFrame, 
            output_path: str,
            main_content_df: Optional[pd.DataFrame] = None,
            page_dfs: Optional[Dict[str, pd.DataFrame]] = None,
            requirement_dfs: Optional[Dict[str, pd.DataFrame]] = None,
            metadata: Optional[Dict[str, Any]] = None
    ) -> None:
        """
        将DataFrame导出为Excel
        
        Args:
            df (pd.DataFrame): 测试用例DataFrame
            output_path (str): 输出文件路径
            main_content_df (Optional[pd.DataFrame]): 主内容区域测试用例DataFrame
            page_dfs (Optional[Dict[str, pd.DataFrame]]): 按页面分组的测试用例DataFrame字典
            requirement_dfs (Optional[Dict[str, pd.DataFrame]]): 按需求文档分组的测试用例DataFrame字典
            metadata (Optional[Dict[str, Any]]): 元数据
        """
        try:
            # 创建一个Excel写入器
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # 导出所有测试用例
                df.to_excel(writer, sheet_name="所有测试用例", index=False)
                self._format_excel_sheet(writer, "所有测试用例", df)
                
                # 如果有主内容区域测试用例，单独导出
                if main_content_df is not None and not main_content_df.empty:
                    main_content_df.to_excel(writer, sheet_name="主内容区域测试用例", index=False)
                    self._format_excel_sheet(writer, "主内容区域测试用例", main_content_df)
                
                # 如果有按页面分组的测试用例，分别导出
                if page_dfs:
                    for page_name, page_df in page_dfs.items():
                        # 处理页面名称，提取域名作为标签
                        if page_name != "多页面" and page_name != "未知页面":
                            sheet_name = f"页面_{extract_domain(page_name)}"
                        else:
                            sheet_name = f"页面_{page_name}"
                            
                        # 确保工作表名称有效（Excel限制工作表名长度为31个字符）
                        sheet_name = sheet_name[:31]
                        page_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        self._format_excel_sheet(writer, sheet_name, page_df)
                
                # 如果有按需求文档分组的测试用例，分别导出
                if requirement_dfs:
                    for req_name, req_df in requirement_dfs.items():
                        if req_name != "综合需求" and req_name != "未知需求":
                            # 移除扩展名
                            base_name = os.path.splitext(req_name)[0]
                            sheet_name = f"需求_{base_name}"
                        else:
                            sheet_name = f"需求_{req_name}"
                            
                        # 确保工作表名称有效
                        sheet_name = sheet_name[:31]
                        req_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        self._format_excel_sheet(writer, sheet_name, req_df)
                
                # 如果有元数据，添加一个元数据工作表
                if metadata:
                    self._add_metadata_sheet(writer, metadata)
            
            logger.info(f"Excel文件已保存到: {output_path}")
            
        except Exception as e:
            logger.error(f"导出DataFrame到Excel时出错: {str(e)}")
            
    def _format_excel_sheet(self, writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
        """
        格式化Excel工作表
        
        Args:
            writer (pd.ExcelWriter): Excel写入器
            sheet_name (str): 工作表名称
            df (pd.DataFrame): 数据DataFrame
        """
        # 获取工作表对象
        worksheet = writer.sheets[sheet_name]
        
        # 设置列宽
        column_widths = {
            "测试ID": 12,
            "测试标题": 30,
            "优先级": 10,
            "前置条件": 20,
            "测试步骤": 40,
            "预期结果": 40,
            "测试数据": 20,
            "页面来源": 25
        }
        
        # 应用列宽
        for i, column in enumerate(df.columns):
            column_letter = get_column_letter(i + 1)
            width = column_widths.get(column, 15)  # 默认宽度为15
            worksheet.column_dimensions[column_letter].width = width
            
        # 格式化标题行
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for i, column in enumerate(df.columns):
            cell = worksheet.cell(row=1, column=i + 1)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            
        # 为单元格应用样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 设置文本对齐和自动换行
        wrap_alignment = Alignment(wrap_text=True, vertical="top")
        
        # 应用样式到所有单元格
        for row in range(1, len(df) + 2):  # +2是因为Excel行从1开始，还有一个标题行
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = wrap_alignment
                
    def _add_metadata_sheet(self, writer: pd.ExcelWriter, metadata: Dict[str, Any]) -> None:
        """
        添加元数据工作表
        
        Args:
            writer (pd.ExcelWriter): Excel写入器
            metadata (Dict[str, Any]): 元数据
        """
        # 创建元数据DataFrame
        metadata_list = []
        
        # 添加时间信息
        metadata_list.append({"项目": "生成时间", "值": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
        
        # 添加测试用例数量
        if "total_cases" in metadata:
            metadata_list.append({"项目": "测试用例总数", "值": metadata["total_cases"]})
            
        # 添加URL信息
        if "urls" in metadata and metadata["urls"]:
            for i, url in enumerate(metadata["urls"]):
                metadata_list.append({"项目": f"测试URL {i+1}", "值": url})
                
        # 添加需求文档信息
        if "requirements" in metadata and metadata["requirements"]:
            for i, req in enumerate(metadata["requirements"]):
                metadata_list.append({"项目": f"需求文档 {i+1}", "值": req})
                
        # 创建DataFrame并导出
        metadata_df = pd.DataFrame(metadata_list)
        metadata_df.to_excel(writer, sheet_name="元数据", index=False)
        
        # 格式化元数据工作表
        self._format_excel_sheet(writer, "元数据", metadata_df) 